import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

def main():
    data_dir = "F:/GoogleDrive_local/1.実験データ_gdrive/5.生データ D/260704 sam 位置合わせ test/foranti"
    output_xlsx = "F:/GoogleDrive_local/1.実験データ_gdrive/5.生データ D/260704 sam 位置合わせ test/foranti/pillar_intensity_analysis_vertical.xlsx"
    
    print("Starting Vertical Excel export pipeline...")
    print(f"Source Folder: {data_dir}")
    print(f"Target Excel:  {output_xlsx}\n")
    
    all_records = []
    
    # 1. foranti フォルダ内のファイルをスキャンして処理
    files = os.listdir(data_dir)
    aligned_files = [f for f in files if f.endswith('_aligned_to_pre.csv')]
    
    # 自然順にソート (1-1, 1-2, ..., 8-8)
    def extract_key(filename):
        base = filename[:-19]
        parts = base.split('-')
        try:
            return (int(parts[0]), int(parts[1]))
        except ValueError:
            return (999, 999)
            
    aligned_files.sort(key=extract_key)
    print(f"Found {len(aligned_files)} aligned CSV files to process.")
    
    total_excluded = 0
    total_matched = 0
    
    for f in aligned_files:
        base = f[:-19] # '1-2'
        parts = base.split('-')
        if len(parts) < 2:
            continue
        cond, set_num = int(parts[0]), int(parts[1])
        
        aligned_path = os.path.join(data_dir, f)
        pre_path = os.path.join(data_dir, f"{base}_pillars_pre.csv")
        
        if not os.path.exists(pre_path):
            print(f"Warning: Pre pillars file not found for {base}. Skipped.")
            continue
            
        # データロード
        df_aligned = pd.read_csv(aligned_path)
        df_pre = pd.read_csv(pre_path)
        
        # 共通部分の抽出 (matched_ref_id != -1)
        matched_df = df_aligned[df_aligned['matched_ref_id'] != -1].copy()
        
        # 除外されたピラー数の計算
        excluded_count = len(df_aligned) - len(matched_df)
        total_excluded += excluded_count
        total_matched += len(matched_df)
        
        # post側の 'pillar_id' をドロップしてマージ時の名前重複を防ぐ
        if 'pillar_id' in matched_df.columns:
            matched_df = matched_df.drop(columns=['pillar_id'])
            
        # preデータとマージして輝度変化を算出
        df_pre_sub = df_pre[['pillar_id', 'box_intensity_3x3']].rename(
            columns={'box_intensity_3x3': 'box_intensity_3x3_pre'}
        )
        
        merged = pd.merge(
            matched_df,
            df_pre_sub,
            left_on='matched_ref_id',
            right_on='pillar_id'
        )
        
        # 輝度変化量 (post - pre) の算出
        merged['intensity_change'] = merged['box_intensity_3x3'] - merged['box_intensity_3x3_pre']
        
        # 必要な列のみを抽出して追加
        temp_df = pd.DataFrame({
            'Condition': cond,
            'Set': set_num,
            'intensity_change': merged['intensity_change']
        })
        
        all_records.append(temp_df)
        
    print(f"\nMerging all data vertically...")
    # 縦にマージ
    final_df = pd.concat(all_records, ignore_index=True)
    print(f"Total matched pillars (rows) to write: {len(final_df):,}")
    print(f"Total excluded pillars: {total_excluded:,}")
    
    # 2. Pandas を用いた Excel への高速書き出し
    print("Writing data to Excel sheet (Vertical format)...")
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        final_df.to_excel(writer, sheet_name='Pillar Intensity', index=False)
        
    # 3. openpyxl を用いたセルのスタイリング (ヘッダー色、グリッド線など)
    print("Applying styling to the Excel sheet...")
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb['Pillar Intensity']
    
    # グリッドラインを表示する設定
    ws.views.sheetView[0].showGridLines = True
    
    # スタイル定義
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Steel Blue
    
    # ヘッダーにスタイル適用 (A1, B1, C1)
    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
    # 列幅の設定
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 25
    
    # 輝度変化量のセル書式を設定 (C列の2行目から最後までを一括処理)
    # 処理負荷低減のため、数式や書式の一括適用
    print("Applying number formatting to intensity column...")
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=3)
        cell.number_format = '0.00'
        
    wb.save(output_xlsx)
    print(f"\nSuccessfully generated Vertical Excel report at:\n{output_xlsx}")

if __name__ == "__main__":
    main()
