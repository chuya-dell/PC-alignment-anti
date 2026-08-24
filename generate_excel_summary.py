import os
import glob
import re
import argparse
import numpy as np
import pandas as pd

def get_bg_and_target_files(input_dir_or_files, bg_prefixes=None):
    """
    Find and group CSV files into background (custom prefixes/numbers or starting with 0) and target files.
    """
    if isinstance(input_dir_or_files, str) and os.path.isdir(input_dir_or_files):
        csv_files = glob.glob(os.path.join(input_dir_or_files, "*.csv"))
    elif isinstance(input_dir_or_files, list):
        csv_files = input_dir_or_files
    else:
        csv_files = glob.glob("*.csv")

    csv_files = sorted(csv_files)
    
    bg_files = []
    tgt_files = []

    def is_bg(filename):
        base = os.path.basename(filename)
        if bg_prefixes:
            str_prefixes = [str(b).strip() for b in bg_prefixes]
            for bg in str_prefixes:
                # Match patterns like "13-", "13_", "No.13", "No13", or starting with "13"
                pattern = rf'(?:^|[\-_bB][gG]?|No\.?)\s*{re.escape(bg)}(?:[\-_\s\.]|$)'
                if re.search(pattern, base, re.IGNORECASE) or base.startswith(bg):
                    return True
            return False
        else:
            match = re.search(r'(\d+)-(\d+)', base)
            if match:
                major, minor = match.group(1), match.group(2)
                if major == '0' or minor == '0':
                    return True
            return base.startswith('0')

    for f in csv_files:
        if is_bg(f):
            bg_files.append(f)
        else:
            tgt_files.append(f)

    return bg_files, tgt_files

def analyze_intensities_excel(input_dir, output_excel_path=None, intensity_col='mean_intensity', sigmas=[3, 5], bg_prefixes=None):
    """
    Compute background mean + n*sigma thresholds and evaluate target sequence CSV files.
    Exports result to a cleanly formatted Excel file (.xlsx).
    """
    if not os.path.exists(input_dir):
        raise FileNotFoundError(f"Input directory does not exist: {input_dir}")

    bg_files, tgt_files = get_bg_and_target_files(input_dir, bg_prefixes=bg_prefixes)

    print(f"=== CSV ファイル検出結果 ===")
    print(f"バックグラウンド (0番台) ファイル数 : {len(bg_files)} 件")
    print(f"測定対象 (ターゲット) ファイル数   : {len(tgt_files)} 件")

    if len(bg_files) == 0:
        print("警告: 頭が0のバックグラウンドファイルが見つかりません。全体の下位20%をバックグラウンド仮定して計算します。")
        all_csvs = sorted(glob.glob(os.path.join(input_dir, "*.csv")))
        if not all_csvs:
            raise FileNotFoundError("指定ディレクトリに CSV ファイルが存在しません。")
        bg_files = [all_csvs[0]]
        tgt_files = all_csvs[1:] if len(all_csvs) > 1 else all_csvs

    # 1. バックグラウンドデータの統計量算出
    bg_intensities = []
    bg_stats_list = []

    for bg_f in bg_files:
        df_bg = pd.read_csv(bg_f)
        col = intensity_col if intensity_col in df_bg.columns else ('intensity' if 'intensity' in df_bg.columns else df_bg.columns[0])
        vals = df_bg[col].dropna().values
        bg_intensities.extend(vals)
        
        bg_stats_list.append({
            'ファイル名': os.path.basename(bg_f),
            'データ種別': 'バックグラウンド (0番台)',
            'データ数 (個)': len(vals),
            '平均輝度 (Mean)': np.mean(vals),
            '輝度標準偏差 (Std)': np.std(vals),
            '中央値 (Median)': np.median(vals),
            '最小値 (Min)': np.min(vals),
            '最大値 (Max)': np.max(vals)
        })

    bg_all_vals = np.array(bg_intensities)
    bg_mean = np.mean(bg_all_vals)
    bg_std = np.std(bg_all_vals)

    print(f"\n--- バックグラウンド基準統計量 ---")
    print(f"統合バックグラウンド平均 (μ_BG) : {bg_mean:.4f}")
    print(f"統合バックグラウンド標準偏差 (σ_BG): {bg_std:.4f}")

    thresholds = {}
    for n in sigmas:
        thresholds[n] = bg_mean + n * bg_std
        print(f"  閾値 (μ + {n}σ): {thresholds[n]:.4f}")

    # 2. 測定対象データの解析
    summary_rows = []

    all_target_files = bg_files + tgt_files
    for tgt_f in all_target_files:
        base_name = os.path.basename(tgt_f)
        is_bg = tgt_f in bg_files
        
        df_tgt = pd.read_csv(tgt_f)
        
        # 位置合わせ(アライメント)済みの評価パターン判定
        is_estimated = bool(df_tgt.get('is_estimated', pd.Series([False]*len(df_tgt))).iloc[0]) if len(df_tgt) > 0 else ("estimated" in base_name.lower())
        has_alignment = 'matched_ref_id' in df_tgt.columns

        if is_estimated or "estimated" in base_name.lower():
            df_active = df_tgt
            filter_note = "Pattern B (全基準格子・推定評価)"
        elif has_alignment:
            df_active = df_tgt[df_tgt['matched_ref_id'] != -1]
            filter_note = "Pattern A (実測同定ペアのみ)"
        else:
            df_active = df_tgt
            filter_note = "単体画像全検出"

        col = intensity_col if intensity_col in df_active.columns else ('intensity' if 'intensity' in df_active.columns else df_active.columns[0])
        vals = df_active[col].dropna().values
        
        total_count = len(vals)
        if total_count == 0:
            continue

        tgt_mean = np.mean(vals)
        tgt_std = np.std(vals)
        delta_mean = tgt_mean - bg_mean

        row = {
            'ファイル名': base_name,
            '分類': 'バックグラウンド' if is_bg else '測定データ',
            '評価パターン': filter_note,
            '規格化母数 [総ピラー数 N_total]': total_count,
            '平均輝度 (μ)': tgt_mean,
            '輝度標準偏差 (σ)': tgt_std,
            'BG比輝度変化 (ΔMean)': delta_mean,
            'BG平均 (μ_BG)': bg_mean,
            'BG標準偏差 (σ_BG)': bg_std
        }

        # 各 n*sigma の超える個数と規格化割合(%)を算出・明確化
        for n in sigmas:
            thresh = thresholds[n]
            count_above = np.sum(vals > thresh)
            ratio_above = count_above / total_count
            row[f'判定閾値 (μ+{n}σ)'] = thresh
            row[f'{n}σ超過数 [個]'] = count_above
            row[f'{n}σ超過規格化割合 [% = (超過数/N_total)*100]'] = ratio_above

        summary_rows.append(row)

    df_summary = pd.DataFrame(summary_rows)
    df_bg_summary = pd.DataFrame(bg_stats_list)

    # 各サンプル系列番号（No.1, No.2 ...）ごとの規格化サマリーシートを作成
    df_summary['サンプル系列'] = df_summary['ファイル名'].apply(lambda x: re.split(r'[_\-\.]', str(x))[0])
    df_series_summary = df_summary.groupby(['サンプル系列', '評価パターン']).agg({
        '規格化母数 [総ピラー数 N_total]': 'mean',
        '平均輝度 (μ)': 'mean',
        'BG比輝度変化 (ΔMean)': 'mean',
        '3σ超過数 [個]': 'mean',
        '3σ超過規格化割合 [% = (超過数/N_total)*100]': 'mean',
        '5σ超過数 [個]': 'mean',
        '5σ超過規格化割合 [% = (超過数/N_total)*100]': 'mean'
    }).reset_index()

    # 3. Excel出力 (.xlsx)
    if output_excel_path is None:
        output_excel_path = os.path.join(input_dir, "intensity_summary_3s_5s.xlsx")

    print(f"\nExcel ファイルに書き込み中: {output_excel_path}")

    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df_summary.drop(columns=['サンプル系列']).to_excel(writer, sheet_name='全データ規格化サマリー', index=False)
        df_series_summary.to_excel(writer, sheet_name='サンプル系列別規格化平均', index=False)
        df_bg_summary.to_excel(writer, sheet_name='背景(BG)基準データ', index=False)

    # 4. openpyxl によるデザイン装飾 (ヘッダー色付け、列幅調整、書式設定)
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.load_workbook(output_excel_path)
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Dark Navy
        header_font = Font(name="Meiryo UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Meiryo UI", size=10)
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            ws.views.sheetView[0].showGridLines = True

            # Header style
            for col_num, cell in enumerate(ws[1], 1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            ws.row_dimensions[1].height = 32

            # Data rows style & format
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.font = data_font
                    cell.border = thin_border
                    col_header = ws.cell(row=1, column=cell.column).value or ""

                    # Number formatting
                    if isinstance(cell.value, (int, float)):
                        if "[%" in col_header or "割合" in col_header:
                            cell.number_format = '0.000%'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif "[個]" in col_header or "N_total" in col_header or "母数" in col_header or "データ数" in col_header:
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        elif any(k in col_header for k in ["平均", "標準偏差", "閾値", "ΔMean", "μ", "σ"]):
                            cell.number_format = '0.0000'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = '#,##0.00'
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            # Auto-fit column widths
            for col in ws.columns:
                max_len = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    val_str = str(cell.value or '')
                    # Handle double-width Japanese characters in width estimation
                    length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                    if length > max_len:
                        max_len = length
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        wb.save(output_excel_path)
        print(f"装飾付き Excel ファイルの出力完了: {output_excel_path}")

    except Exception as e:
        print(f"openpyxl 装飾中の注意 (ファイルは書き出されています): {e}")

    return output_excel_path

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Generate Excel summary comparing sequence data against 0-series background Mean + n*sigma")
    parser.add_argument("--dir", type=str, required=True, help="Directory containing CSV sequence files (0-1.csv, 1-1.csv, etc.)")
    parser.add_argument("--out", type=str, default=None, help="Output Excel (.xlsx) file path")
    parser.add_argument("--col", type=str, default="mean_intensity", help="Column name for intensity data (default: mean_intensity)")
    parser.add_argument("--sigmas", type=int, nargs="+", default=[3, 5], help="List of n values for mean + n*sigma (default: 3 5)")
    
    args = parser.parse_args()
    analyze_intensities_excel(args.dir, args.out, args.col, sigmas=args.sigmas)
