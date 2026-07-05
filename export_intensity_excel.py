import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import time
import argparse

def main():
    parser = argparse.ArgumentParser(description="Pillar Intensity Excel Aggregator")
    parser.add_argument("data_dir", nargs="?", 
                        default="F:/GoogleDrive_local/1.実験データ_gdrive/5.生データ D/260704 sam 位置合わせ test/foranti", 
                        help="Path to the directory containing aligned CSV files")
    parser.add_argument("--output-local", action="store_true", 
                        help="Also output to local git repository as ./pillar_intensity_analysis_120k.xlsx")
    args = parser.parse_args()
    
    # パス解決
    data_dir = args.data_dir.replace('\\', '/')
    output_xlsx = os.path.join(data_dir, "pillar_intensity_analysis.xlsx").replace('\\', '/')
    
    print("Starting Optimized Excel export pipeline (1-120k index in col A)...")
    print(f"Source Folder: {data_dir}")
    print(f"Target Excel:  {output_xlsx}")
    if args.output_local:
        print("Target Local:  ./pillar_intensity_analysis_120k.xlsx")
    print()
    
    if not os.path.exists(data_dir):
        print(f"Error: Source folder does not exist: {data_dir}")
        return
        
    # 8つの条件 (1〜8) ごとにデータを保持
    cond_sheets = {str(c): {} for c in range(1, 9)}
    
    # 1. foranti フォルダ内のファイルをスキャンして処理
    files = os.listdir(data_dir)
    aligned_files = [f for f in files if f.endswith('_aligned_to_pre.csv')]
    
    # 自然順にソート (1-1, 1-2, ..., 8-8)
    def extract_key(filename):
        base = filename[:-19]
        parts = base.split('-')
        try:
            return (int(parts[0]), int(parts[1]))
        except ValueError:
            return (999, 999)
            
    aligned_files.sort(key=extract_key)
    print(f"Found {len(aligned_files)} aligned CSV files to process.")
    
    for f in aligned_files:
        base = f[:-19] # '1-2'
        parts = base.split('-')
        if len(parts) < 2:
            continue
        cond, set_num = parts[0], parts[1]
        
        if cond not in cond_sheets:
            continue
            
        aligned_csv = os.path.join(data_dir, f)
        post_csv = os.path.join(data_dir, f"{base}_pillars_post.csv")
        pre_csv = os.path.join(data_dir, f"{base}_pillars_pre.csv")
        
        if not os.path.exists(post_csv) or not os.path.exists(pre_csv):
            print(f"Warning: Original CSVs not found for {base}. Skipped.")
            continue
            
        df_aligned = pd.read_csv(aligned_csv)
        df_post_orig = pd.read_csv(post_csv)
        df_pre = pd.read_csv(pre_csv)
        
        # 傷部分の除外処理 (ICPマッチ率による自動外れ値フィルタリング)
        # ICPでマッチしなかった (matched_ref_id == -1) ピラーのうち、
        # 傷やゴミなどによる異常高輝度のもの (box_intensity_3x3 > 120) を検出し、
        # アライメントから外れたゴミピラーとして計数・除外する
        unmatched_tgt = df_aligned[df_aligned['matched_ref_id'] == -1]
        excluded_count = len(unmatched_tgt[unmatched_tgt['box_intensity_3x3'] > 120])
        
        # 正常にマッチした (matched_ref_id != -1) ピラーのみを抽出
        matched_df = df_aligned[df_aligned['matched_ref_id'] != -1].copy()
        
        # 基準側 (pre) ピラーの輝度をマージして取得 (衝突防止のため pre_pillar_id にリネーム)
        df_pre_sub = df_pre[['pillar_id', 'box_intensity_3x3']].rename(
            columns={'pillar_id': 'pre_pillar_id', 'box_intensity_3x3': 'box_intensity_3x3_pre'}
        )
        
        merged = pd.merge(
            matched_df,
            df_pre_sub,
            left_on='matched_ref_id',
            right_on='pre_pillar_id'
        )
        
        # 輝度変化量 (post - pre) の算出
        merged['intensity_change'] = merged['box_intensity_3x3'] - merged['box_intensity_3x3_pre']
        
        # 出力データの整理 (基準側IDと輝度変化量)
        result_df = merged[['pre_pillar_id', 'intensity_change']].rename(columns={'pre_pillar_id': 'pillar_id'}).copy()
        result_df = result_df.sort_values('pillar_id').reset_index(drop=True)
        
        cond_sheets[cond][set_num] = {
            'total_count': len(result_df),
            'excluded_count': excluded_count,
            'data': result_df
        }
        
    # 2. Pandas を使って 120,000行の基礎 DataFrame を一気に Excel へ高速書き出し (リトライ機能付き)
    print("\nWriting main data columns (A-I) to Excel sheets via Pandas...")
    
    writer = None
    for attempt in range(1, 21):
        try:
            writer = pd.ExcelWriter(output_xlsx, engine='openpyxl')
            break
        except PermissionError:
            print(f"Warning: Excel file is locked (Permission denied). Attempt {attempt}/20. Please close the Excel file if it is open...")
            time.sleep(3)
            
    if writer is None:
        raise PermissionError(f"Could not open Excel file for writing because it is locked: {output_xlsx}")
        
    with writer:
        for cond_num in range(1, 9):
            cond_str = str(cond_num)
            sheet_name = f"Condition {cond_str}"
            
            # A列: インデックス (1〜120,000)
            df_cond = pd.DataFrame({'': np.arange(1, 120001)})
            
            sets_data = cond_sheets[cond_str]
            # B〜I列: セット1〜8 of 輝度変化量 (1〜12万インデックスに対応)
            for s_idx in range(1, 9):
                set_str = str(s_idx)
                col_name = f"Set {set_str}"
                
                if set_str in sets_data:
                    # 輝度変化データ
                    values = sets_data[set_str]['data']['intensity_change'].values
                    # 120,000行に引き伸ばして足りない部分を NaN (空白) で埋める
                    padded_values = np.full(120000, np.nan)
                    padded_values[:len(values)] = values
                    df_cond[col_name] = padded_values
                else:
                    df_cond[col_name] = np.full(120000, np.nan)
                    
            # インデックス無し、ヘッダーもそのまま書き出し
            df_cond.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # 3. openpyxl を使って、K列・L列の統計情報テーブルの追加と書式スタイリングを行う
    print("Applying styling and adding Set statistics to K-L columns...")
    
    wb = None
    for attempt in range(1, 21):
        try:
            wb = openpyxl.load_workbook(output_xlsx)
            break
        except PermissionError:
            print(f"Warning: Excel file is locked during post-styling open. Attempt {attempt}/20. Retrying...")
            time.sleep(3)
            
    if wb is None:
        raise PermissionError(f"Could not read Excel file because it is locked by Excel: {output_xlsx}")
    
    header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
    normal_font = Font(name="Segoe UI", size=9)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Steel Blue
    
    for cond_num in range(1, 9):
        cond_str = str(cond_num)
        sheet_name = f"Condition {cond_str}"
        ws = wb[sheet_name]
        
        # グリッドラインを明示的に有効化
        ws.views.sheetView[0].showGridLines = True
        
        # 1行目はデータから直接始まる仕様のため、ヘッダー行をデータの実数値で上書き
        # A1セル: 連番スタート値「1」
        c_a = ws.cell(row=1, column=1)
        c_a.value = 1
        c_a.alignment = Alignment(horizontal="center")
        c_a.font = normal_font
        
        # B1〜I1セル: 各セットデータの最初の値
        sets_data = cond_sheets[cond_str]
        for s_idx in range(1, 9):
            set_str = str(s_idx)
            col_idx = s_idx + 1 # B (2) 〜 I (9)
            c_data = ws.cell(row=1, column=col_idx)
            
            if set_str in sets_data and len(sets_data[set_str]['data']) > 0:
                first_val = float(sets_data[set_str]['data']['intensity_change'].iloc[0])
                if not np.isnan(first_val):
                    c_data.value = first_val
                    c_data.number_format = '0.00'
                else:
                    c_data.value = None
            else:
                c_data.value = None
                
            c_data.alignment = Alignment(horizontal="right")
            c_data.font = normal_font
            
        # J1: 空白
        ws.cell(row=1, column=10).value = None
        
        # K1: "Set", L1: "N" でヘッダーを上書き
        c_k_header = ws.cell(row=1, column=11, value="Set")
        c_l_header = ws.cell(row=1, column=12, value="N")
        
        c_k_header.font = header_font
        c_l_header.font = header_font
        c_k_header.fill = header_fill
        c_l_header.fill = header_fill
        c_k_header.alignment = Alignment(horizontal="center")
        c_l_header.alignment = Alignment(horizontal="center")
        
        # K2〜K9: セット名、L2〜L9: 各セットの有効ピラー数 (N数)
        for s_idx in range(1, 9):
            set_str = str(s_idx)
            row_info = s_idx + 1
            
            c_k = ws.cell(row=row_info, column=11, value=f"Set {set_str}")
            c_k.alignment = Alignment(horizontal="center")
            
            if set_str in sets_data:
                valid_count = sets_data[set_str]['total_count']
                c_l = ws.cell(row=row_info, column=12, value=valid_count)
                c_l.alignment = Alignment(horizontal="right")
                c_l.number_format = '#,##0'
            else:
                c_l = ws.cell(row=row_info, column=12, value="No Data")
                c_l.alignment = Alignment(horizontal="center")
                
            c_k.font = normal_font
            c_l.font = normal_font
            c_k.alignment = Alignment(horizontal="center")
            
        # 列幅設定
        ws.column_dimensions['A'].width = 10
        for s_idx in range(1, 9):
            col_letter = get_column_letter(s_idx + 1)
            ws.column_dimensions[col_letter].width = 12
            
        ws.column_dimensions['J'].width = 3 # 空き列
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 15
        
    # 保存処理 (Google Drive と Local の両方に書き出す、ファイルロックに備えてリトライ付き)
    out_paths = [output_xlsx]
    if args.output_local:
        out_paths.append("./pillar_intensity_analysis_120k.xlsx")
        
    for out_path in out_paths:
        print(f"Saving workbook to: {out_path}")
        for attempt in range(1, 21):
            try:
                # 親ディレクトリ作成 (必要な場合)
                os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
                wb.save(out_path)
                print(f"  Successfully saved to: {out_path}")
                break
            except PermissionError:
                print(f"  Warning: File is locked. Attempt {attempt}/20. Retrying in 3s...")
                time.sleep(3)

if __name__ == "__main__":
    main()
