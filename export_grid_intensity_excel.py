import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.utils import get_column_letter
import argparse

def main():
    parser = argparse.ArgumentParser(description="Grid Intensity Excel Aggregator")
    parser.add_argument("data_dir", nargs="?", 
                        default="F:/GoogleDrive_local/1.実験データ_gdrive/5.生データ D/260704 sam 位置合わせ test/foranti/grid_analysis", 
                        help="Path to the directory containing grid_analysis CSV files")
    parser.add_argument("--output-local", action="store_true", 
                        help="Also output to local git repository as ./grid_intensity_analysis.xlsx")
    args = parser.parse_args()
    
    # バックスラッシュをスラッシュに置換して統一
    data_dir = args.data_dir.replace('\\', '/')
    output_xlsx = os.path.join(data_dir, "grid_intensity_analysis.xlsx").replace('\\', '/')
    
    print("Starting Grid Intensity Excel Aggregation...")
    print(f"Source Folder: {data_dir}")
    print(f"Target Excel:  {output_xlsx}")
    if args.output_local:
        print("Target Local:  ./grid_intensity_analysis.xlsx")
    print()
    
    if not os.path.exists(data_dir):
        print(f"Error: Source folder does not exist: {data_dir}")
        return
        
    # グリッドの解像度情報 (2048x2044, G=6.29)
    # cols = 325, rows = 324, total = 105,300
    total_grids = 105300
    expected_cols = 325
    expected_rows = 324
    
    # 事前にK列・L列（Row, Col座標）のデータを1回生成しておく
    rows_coords = []
    cols_coords = []
    for r in range(1, expected_rows + 1):
        for c in range(1, expected_cols + 1):
            rows_coords.append(r)
            cols_coords.append(c)
            
    # 8つの条件 (1〜8) ごとにデータを保持
    cond_sheets = {str(c): {} for c in range(1, 9)}
    
    # 1. grid_analysis フォルダ内のファイルをスキャン
    files = os.listdir(data_dir)
    grid_files = [f for f in files if f.endswith('_grid_analysis.csv')]
    
    for f in grid_files:
        base = f[:-18] # '1-1'
        parts = base.split('-')
        if len(parts) < 2:
            continue
        cond, set_num = parts[0], parts[1]
        
        if cond not in cond_sheets:
            continue
            
        csv_path = os.path.join(data_dir, f)
        df = pd.read_csv(csv_path)
        
        # 整合性を保つため、Row -> Col 順でソート
        df_sorted = df.sort_values(by=['Row', 'Col']).reset_index(drop=True)
        
        cond_sheets[cond][set_num] = df_sorted
        
    # 2. Pandas ExcelWriter を用いて高速書き出し
    print("Writing main columns to Excel via Pandas...")
    with pd.ExcelWriter(output_xlsx, engine='openpyxl') as writer:
        for cond_num in range(1, 9):
            cond_str = str(cond_num)
            sheet_name = f"Condition {cond_str}"
            
            # A列: インデックス (1〜105,300)
            df_cond = pd.DataFrame({'': np.arange(1, total_grids + 1)})
            
            sets_data = cond_sheets[cond_str]
            # B〜I列: セット1〜8 of Mean データを流し込む
            for s_idx in range(1, 9):
                set_str = str(s_idx)
                col_name = f"Set {set_str}"
                
                if set_str in sets_data:
                    mean_vals = sets_data[set_str]['Mean'].values
                    # 105,300行に引き伸ばして足りない部分は NaN (空白) で埋める
                    padded_vals = np.full(total_grids, np.nan)
                    padded_vals[:len(mean_vals)] = mean_vals
                    df_cond[col_name] = padded_vals
                else:
                    df_cond[col_name] = np.full(total_grids, np.nan)
                    
            # J列(10番目の列)は空白列とする
            df_cond[' '] = np.full(total_grids, np.nan)
            
            # K列, L列: Row, Col 座標
            df_cond['Row'] = rows_coords
            df_cond['Col'] = cols_coords
            
            # インデックス無し、ヘッダーもそのまま書き出し
            df_cond.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # 3. openpyxl を使って、1行目のヘッダーセルの上書きと表示調整
    print("Adjusting row 1 and column styles in openpyxl...")
    wb = openpyxl.load_workbook(output_xlsx)
    
    normal_font = openpyxl.styles.Font(name="Segoe UI", size=9)
    align_right = openpyxl.styles.Alignment(horizontal="right")
    align_center = openpyxl.styles.Alignment(horizontal="center")
    
    for cond_num in range(1, 9):
        cond_str = str(cond_num)
        sheet_name = f"Condition {cond_str}"
        ws = wb[sheet_name]
        
        # グリッドラインを表示する設定 (Excelで枠線を表示させる)
        ws.views.sheetView[0].showGridLines = True
        
        # 1行目はすべて数値データから始まる仕様 (ヘッダー行をデータの実数値で上書き)
        # A1セル: 連番スタート値「1」
        cell_a = ws.cell(row=1, column=1)
        cell_a.value = 1
        cell_a.font = normal_font
        cell_a.alignment = align_center
        
        # B1〜I1セル: 各セットデータの最初の値
        sets_data = cond_sheets[cond_str]
        for s_idx in range(1, 9):
            set_str = str(s_idx)
            col_idx = s_idx + 1 # B (2) 〜 I (9)
            cell_data = ws.cell(row=1, column=col_idx)
            
            if set_str in sets_data and len(sets_data[set_str]) > 0:
                first_val = float(sets_data[set_str]['Mean'].iloc[0])
                if not np.isnan(first_val):
                    cell_data.value = first_val
                    cell_data.number_format = '0.00'
                else:
                    cell_data.value = None
            else:
                cell_data.value = None
                
            cell_data.font = normal_font
            cell_data.alignment = align_right
            
        # J1セル: 空白
        ws.cell(row=1, column=10).value = None
        
        # K1セル: Rowの1点目「1」
        cell_k = ws.cell(row=1, column=11)
        cell_k.value = 1
        cell_k.font = normal_font
        cell_k.alignment = align_right
        
        # L1セル: Colの1点目「1」
        cell_l = ws.cell(row=1, column=12)
        cell_l.value = 1
        cell_l.font = normal_font
        cell_l.alignment = align_right
        
        # 列幅の設定
        ws.column_dimensions['A'].width = 10
        for s_idx in range(1, 9):
            col_letter = get_column_letter(s_idx + 1)
            ws.column_dimensions[col_letter].width = 12
            
        ws.column_dimensions['J'].width = 3 # 空き列
        ws.column_dimensions['K'].width = 10
        ws.column_dimensions['L'].width = 10
        ws.column_dimensions['M'].width = 3 # 空き列
        ws.column_dimensions['N'].width = 10 # 統計Set名
        ws.column_dimensions['O'].width = 15 # 有効グリッド数
        
        # N列・O列に各セットの有効グリッド数を縦一列で書き込む
        for s_idx in range(1, 9):
            set_str = str(s_idx)
            valid_count = 0
            if set_str in sets_data:
                valid_count = int(sets_data[set_str]['Mean'].notna().sum())
            
            # N列: Set名
            cell_name = ws.cell(row=s_idx, column=14, value=f"Set {set_str}")
            cell_name.font = normal_font
            
            # O列: 有効グリッド数
            cell_val = ws.cell(row=s_idx, column=15, value=valid_count)
            cell_val.font = normal_font
            cell_val.alignment = align_right
        
    # 保存処理 (ファイルロックに備えてリトライ付き)
    out_paths = [output_xlsx]
    if args.output_local:
        out_paths.append("./grid_intensity_analysis.xlsx")
        
    import time
    for out_path in out_paths:
        print(f"Saving workbook to: {out_path}")
        for attempt in range(1, 21):
            try:
                # 親フォルダ作成
                os.makedirs(os.path.dirname(os.path.abspath(out_path)), exist_ok=True)
                wb.save(out_path)
                print(f"  Successfully saved to: {out_path}")
                break
            except PermissionError:
                print(f"  Warning: File is locked. Attempt {attempt}/20. Retrying in 3s...")
                time.sleep(3)

if __name__ == "__main__":
    main()
